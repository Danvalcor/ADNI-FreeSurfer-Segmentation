import os
import subprocess
import numpy as np
import pandas as pd
import nibabel as nib
from openpyxl import Workbook


# Retrieves the freesurfer installation path and executes the setup file.
freeSurfer = os.environ['FREESURFER_HOME']
subprocess.run([f"{freeSurfer}/SetUpFreeSurfer.sh"])

## Path Directories

"""
    Default Sctructure to follow:
    # Main path where all requiered files are encountered.
    mainPath = os.path.join('/home/danvalcor/Documents/ANDI/Sample010')

    # Specific path for the volumes.
    asegPath1 = f'{mainPath}/mri/aparc+aseg.mgz'
    asegPath2 = f'{mainPath}/mri/aparc+aseg.mgz'
"""

# Main path where all requiered files are encountered.
mainPath = os.path.join('/home/danvalcor/Downloads/ADNI')

# Specific path for the volumes.

# ADNI Segmentation.
asegPaths1 = [f'{mainPath}/I13722/I167591/aparc+aseg.mgz', 
              f'{mainPath}/I28561/I178297/aparc+aseg.mgz',
              f'{mainPath}/I55275/I177649/aparc+aseg.mgz',
              f'{mainPath}/I55771/CRI55771/I178302/aparc+aseg.mgz',
              f'{mainPath}/I114210/I177659/aparc+aseg.mgz',
              f'{mainPath}/I144446/I185652/aparc+aseg.mgz',
              f'{mainPath}/I29705/CRI29705/I178300/aparc+aseg.mgz',
              f'{mainPath}/I14437/CRI14437/CRI14437/aparc+aseg.mgz']          

# FreeSurfer Segmentation.
asegPaths2 = [f'{mainPath}/I13722/PI13722/mri/aparc+aseg.mgz', 
              f'{mainPath}/I28561/PI28561/mri/aparc+aseg.mgz',
              f'{mainPath}/I55275/PI55275/mri/aparc+aseg.mgz',
              f'{mainPath}/I55771/PI55771/mri/aparc+aseg.mgz',
              f'{mainPath}/I114210/PI114210/mri/aparc+aseg.mgz',
              f'{mainPath}/I144446/PI144446/mri/aparc+aseg.mgz',
              f'{mainPath}/I29705/PI29705/mri/aparc+aseg.mgz',
              f'{mainPath}/I14437/PI14437/mri/aparc+aseg.mgz']      

## Lut Table.  
# Path to load the LUT table
LUTfile = os.path.join(os.environ['FREESURFER_HOME'],'luts/FreeSurferColorLUT.txt')

# Declares the path where the LUT file is stored inside FreeSurfers Directory. 
lut_data = []

# Opens and reads the file
with open(LUTfile, "r") as file:
        content = file.readlines()

# Reads each line of the file
for line in content:
    # Skip empty lines
    if len(line) > 1:
        # Strip, split and join each line by the tab ('\t') separator, to assign each value easier
        sLine = line.strip().split("\t")
        sLine = " ".join(sLine).split()
        # Checks and skipps anotations in the file
        if '#' not in sLine[0]:
            # Separates and assigns the contained values to the dictionary
            labelId = int(sLine[0])                     # Converts the numbers back into integers.
            labelName = sLine[1]                        # Assigns the structure labels.
            rgbColors = [int(p) for p in sLine[-4:-1]]  # Converts each rgb value back into integers.
            lut_data.append({"Id": labelId, "StructName": labelName, "Color Array": rgbColors})

# Create DataFrame from the list of dictionaries
lut = pd.DataFrame(lut_data)# Prints the dictionary to verify

# Creates an Excel Workbook
wb = Workbook()

# Gets the active sheet.
ws = wb.active

for i, (asegPath1, asegPath2) in enumerate(zip(asegPaths1, asegPaths2)):
    df = lut.copy(deep=True) # Creates a copy from the LUT table. 
    
    # Gets the subject Id from the route. May need to modify according to personal needs.
    parts = asegPath2.split('/')
    subjectID = None
    for p in parts:
        if p.startswith('I'):
            subjectID = p
            break
    
    
    ## Filtering Segments from both segmentations.
    # Load the MRI libraries using the nibabel library
    aseg1 = nib.load(asegPath1)     # Volume containing Freesurfer Segmentation. 
    aseg2 = nib.load(asegPath2)     # Volume containing ADNI Segmentation. 

    # We obtain and save the data present in our segmentation volume.
    img_data1 = aseg1.get_fdata()
    img_data2 = aseg2.get_fdata()

    # Create a temporary array to store the number of voxels present in every segment.
    calculated1 = []
    calculated2 = []

    # Iterate over each row present in the DataFrame
    for _, row in df.iterrows():
        # Assign the id, structure Name, and corresponding color
        colorId, name, color = row[['Id', 'StructName', 'Color Array']]
        # Create a boolean mask for values in img_data matching the Id value
        mask = (img_data1 == colorId)
        num_voxels = np.count_nonzero(mask)
        calculated1.append(num_voxels)
        mask = (img_data2 == colorId)
        num_voxels = np.count_nonzero(mask)
        calculated2.append(num_voxels)

    # Add the calculated lists as new columns named "NVoxels1" and "NVoxels2" to the DataFrame
    df['NVoxels1'] = calculated1
    df['NVoxels2'] = calculated2

    # Remove rows where 'NVoxels' is 0, and reset the index of the DataFrame
    df = df[(df['NVoxels1'] != 0) & (df['NVoxels2'] != 0)].reset_index(drop=True)

    # Create a column with the difference of voxels.
    df['DiffVoxels'] = abs(df['NVoxels1'] - df['NVoxels2'])

    # Show the resulting DataFrame
    print("Present Segments: ",len(df))
    difSeg = len(df[df['DiffVoxels']>0])
    if difSeg>0:
        print("Different Segments: ",difSeg)
        print(df[["StructName","NVoxels1","NVoxels2",'DiffVoxels']][df['DiffVoxels']>0])
    else:
        print("No found difference in the segments: ")
        print(df[["StructName","NVoxels1","NVoxels2",'DiffVoxels']])

    # Selects desired columns to store on the worksheet.
    colums = ["Id","StructName", "NVoxels1", "NVoxels2", 'DiffVoxels']
    dfCols = df[colums]

    # Obtaines the rows from the dataframe.
    rows = dfCols.values.tolist()

    # Creates a new sheet for each subject.
    if i == 0: # For the active sheet, assign the first subject Id.
        ws.title = subjectID
    else:
        ws = wb.create_sheet(title=subjectID)

    # Write the DF headers
    encabezados = colums
    for col, encabezado in enumerate(encabezados, start=1):
        ws.cell(row=1, column=col, value=encabezado)

    # Write the desired information in the corresponding sheet.
    for r_idx, fila in enumerate(rows, start=2): 
        for c_idx, valor in enumerate(fila, start=1): 
            ws.cell(row=r_idx, column=c_idx, value=valor)

# Saves the SpreadSheet.
wb.save("comparison_data.xlsx")