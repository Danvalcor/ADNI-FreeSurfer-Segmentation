import os
import numpy as np
import pandas as pd
import nibabel as nib
from helperFuns import *
from openpyxl import Workbook

## Path Directories

# Main path where all requiered files are encountered.
mainPath = os.path.join('/home/danvalcor/Downloads/ADNI/I13722')

# Specific path for the volumes.
asegPath1 = f'{mainPath}/I167591/aparc+aseg.mgz' # ADNI Segmentation.
asegPath2 = f'{mainPath}/PI13722/mri/aparc+aseg.mgz' # FreeSurfer Segmentation.

## Load lut Table.  
df = loadLUT()

# Creates an Excel Workbook and gets the active sheet.
wb = Workbook()
ws = wb.active

# Gets the subject Id from the route. May need to modify according to personal needs.
parts = asegPath2.split('/')
subjectID = None
for p in parts:
    if p.startswith('I'):
        subjectID = p
        break
    
## Filtering Segments from both segmentations.
# Load the MRI libraries using the nibabel library
aseg1 = nib.load(asegPath1)     # Volume containing ADNI Segmentation.
aseg2 = nib.load(asegPath2)     # Volume containing Freesurfer Segmentation. 

dfCols = compareADNI(adni=aseg1, freesurfer=aseg2)

# Obtaines the rows from the dataframe.
rows = dfCols.values.tolist()

# For the active sheet, assign the first subject Id.
ws.title = subjectID

# Write the DF headers
headers  = dfCols.columns.to_numpy()

for col, header in enumerate(headers, start=1):
    ws.cell(row=1, column=col, value=header)

# Write the desired information in the corresponding sheet.
for r_idx, fila in enumerate(rows, start=2): 
    for c_idx, valor in enumerate(fila, start=1): 
        ws.cell(row=r_idx, column=c_idx, value=valor)

# Stats folder path
statsPath = f'{mainPath}/stats'

# Verify path
if not os.path.exists(statsPath):
    # Si la carpeta no existe, crearla
    os.makedirs(statsPath)

# Saves the SpreadSheet.
wb.save(f'{statsPath}/comparison_data.xlsx')