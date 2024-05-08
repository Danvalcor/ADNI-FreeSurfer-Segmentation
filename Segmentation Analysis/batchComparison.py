import os
import subprocess
import nibabel as nib
from helperFuns import *
from openpyxl import Workbook


# Retrieves the freesurfer installation path and executes the setup file.
freeSurfer = os.environ['FREESURFER_HOME']
subprocess.run([f"{freeSurfer}/SetUpFreeSurfer.sh"])

# Main path where all requiered files are encountered.
mainPath = os.path.join('/home/danvalcor/Downloads/ADNI')

# Specific path for the volumes.

# ADNI Segmentation.
asegPaths1 = [f'{mainPath}/I13722/I167591/aparc+aseg.mgz', 
              f'{mainPath}/I28561/I178297/aparc+aseg.mgz',
              f'{mainPath}/I55275/I177649/aparc+aseg.mgz',
              f'{mainPath}/I55771/CRI55771/I178302/aparc+aseg.mgz',
              f'{mainPath}/I114210/I177659/aparc+aseg.mgz',
              f'{mainPath}/I144446/I185652/aparc+aseg.mgz',
              f'{mainPath}/I29705/CRI29705/I178300/aparc+aseg.mgz',
              f'{mainPath}/I14437/CRI14437/CRI14437/aparc+aseg.mgz']          

# FreeSurfer Segmentation.
asegPaths2 = [f'{mainPath}/I13722/PI13722/mri/aparc+aseg.mgz', 
              f'{mainPath}/I28561/PI28561/mri/aparc+aseg.mgz',
              f'{mainPath}/I55275/PI55275/mri/aparc+aseg.mgz',
              f'{mainPath}/I55771/PI55771/mri/aparc+aseg.mgz',
              f'{mainPath}/I114210/PI114210/mri/aparc+aseg.mgz',
              f'{mainPath}/I144446/PI144446/mri/aparc+aseg.mgz',
              f'{mainPath}/I29705/PI29705/mri/aparc+aseg.mgz',
              f'{mainPath}/I14437/PI14437/mri/aparc+aseg.mgz']      

## Load lut Table.  
lut = loadLUT()

# Creates an Excel Workbook
wb = Workbook()

# Gets the active sheet.
ws = wb.active

for i, (asegPath1, asegPath2) in enumerate(zip(asegPaths1, asegPaths2)):
    df = lut.copy(deep=True) # Creates a copy from the LUT table. 
    
    # Gets the subject Id from the route. May need to modify according to personal needs.
    parts = asegPath2.split('/')
    subjectID = None
    for p in parts:
        if p.startswith('I'):
            subjectID = p
            break
    
    
    ## Filtering Segments from both segmentations.
    # Load the MRI libraries using the nibabel library
    aseg1 = nib.load(asegPath1)     # Volume containing ADNI Segmentation.
    aseg2 = nib.load(asegPath2)     # Volume containing Freesurfer Segmentation. 

    dfCols = comparison(adni=aseg1, freesurfer=aseg2)

    # Obtaines the rows from the dataframe.
    rows = dfCols.values.tolist()

    # Creates a new sheet for each subject.
    if i == 0: # For the active sheet, assign the first subject Id.
        ws.title = subjectID
    else:
        ws = wb.create_sheet(title=subjectID)

    # Write the DF headers
    headers  = dfCols.columns.to_numpy()
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    # Write the desired information in the corresponding sheet.
    for r_idx, fila in enumerate(rows, start=2): 
        for c_idx, valor in enumerate(fila, start=1): 
            ws.cell(row=r_idx, column=c_idx, value=valor)

# Saves the SpreadSheet.
wb.save("batch_comparison_data.xlsx")